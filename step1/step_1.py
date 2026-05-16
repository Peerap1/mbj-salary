# python3 step1/step1.py

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# --- กำหนดไฟล์และ sheet ---
input_file = "data/1-15 พฤษภาคม 2569.xls"
output_file = "excel/highlight.xlsx"
sheet_name = "Sheet1"

# อ่านไฟล์ .xls
df = pd.read_excel(input_file, sheet_name=None)  # sheet_name=None เพื่ออ่านทุกชีต

# สร้างไฟล์ .xlsx ใหม่
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    for sheet_name, sheet_df in df.items():
        sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"✅ แปลงสำเร็จ: {output_file}")

# --- อ่าน Excel ด้วย pandas ---
df = pd.read_excel(output_file, sheet_name=sheet_name)

# แปลง 'วัน/เวลา' เป็น datetime
df.columns = df.columns.str.strip()
df['วัน/เวลา'] = pd.to_datetime(df['วัน/เวลา'], format='%d/%m/%Y %H:%M:%S')

# เพิ่มคอลัมน์ 'วันที่' เพื่อจัดกลุ่ม
df['วันที่'] = df['วัน/เวลา'].dt.date

# โหลด workbook ด้วย openpyxl เพื่อไฮไลท์
wb = load_workbook(output_file)
ws = wb[sheet_name]

# นิยามสีไฮไลท์
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")       # แดง
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")    # เหลือง

# จัดกลุ่มตามชื่อ-นามสกุล และวันที่
grouped = df.groupby(['ชื่อ-นามสกุล', 'วันที่'])

# วนลูปตรวจสอบแต่ละกลุ่ม
for (name, date), group in grouped:
    count = len(group)
    if count < 4:
        fill = red_fill
    elif count > 4:
        fill = yellow_fill
    else:
        continue  # ถ้าจำนวนแถว = 4 ไม่ต้องไฮไลท์

    for idx in group.index:
        excel_row = idx + 2  # +2 เพราะแถวใน Excel เริ่มที่ 1 และมี header
        for col in range(1, 4):  # คอลัมน์ A (1), B (2), C (3)
            ws.cell(row=excel_row, column=col).fill = fill

# --- บันทึกไฟล์ใหม่ ---
wb.save(output_file)
print(f"✅ บันทึกไฟล์ไฮไลท์แล้วที่: {output_file}")