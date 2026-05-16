import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import yaml

# === STEP 0: โหลด config ===
with open("configs/pay.yaml", "r", encoding="utf-8") as f:
    data = yaml.safe_load(f)

input_file = "data/highlight 1-15 พฤษภาคม 2569 (1).xlsx"
pre_output = "from_highlight"
sheet = "Sheet1"

# === STEP 1: อ่านไฟล์ Excel และเตรียมข้อมูล ===
df = pd.read_excel(input_file, sheet_name=sheet)
df.columns = df.columns.str.strip()
df['วัน/เวลา'] = pd.to_datetime(df['วัน/เวลา'], format='%d/%m/%Y %H:%M:%S')
df['วันที่'] = df['วัน/เวลา'].dt.date

# เพิ่มคอลัมน์ใหม่
df['ชื่อ'] = ""
df['เวลารวม'] = ""
df['ชื่อ(รวม)'] = ""
df['เวลารวมทั้งหมด'] = ""
df['เวลากดแผ่น'] = ""
df['เวลาห้องแพ็ค'] = ""

# === STEP 2: คำนวณเวลารวมรายวัน ===
grouped_daily = df.groupby(['รหัสที่เครื่อง', 'ชื่อ-นามสกุล', 'วันที่'])

for (code, name, date), group in grouped_daily:
    group_sorted = group.sort_values('วัน/เวลา')
    times = group_sorted['วัน/เวลา'].tolist()
    total_minutes = 0
    for i in range(1, len(times), 2):
        total_minutes += (times[i] - times[i - 1]).total_seconds() / 60
    hours, minutes = divmod(int(total_minutes), 60)
    df.at[group_sorted.index[0], 'ชื่อ'] = name
    df.at[group_sorted.index[0], 'เวลารวม'] = f"{hours:02}:{minutes:02}"

# === STEP 3: คำนวณเวลารวมทั้งหมด และแยกเวลากดแผ่น/ห้องแพ็ค ===
target_people = data['people_kod']
grouped_person = df.groupby(['รหัสที่เครื่อง', 'ชื่อ-นามสกุล'])

for (code, name), group in grouped_person:
    name_str = str(name).strip()
    total_minutes_all = 0
    press_minutes = 0
    daily_group = group.groupby(group['วัน/เวลา'].dt.date)

    for date, day_group in daily_group:
        times = sorted(day_group['วัน/เวลา'].tolist())
        if any(p in name_str for p in target_people) and len(times) >= 2:
            first_time = times[0].time()
            if datetime.strptime("00:01:00", "%H:%M:%S").time() <= first_time <= datetime.strptime("02:50:00", "%H:%M:%S").time():
                press_minutes += (times[1] - times[0]).total_seconds() / 60

        for i in range(1, len(times), 2):
            total_minutes_all += (times[i] - times[i - 1]).total_seconds() / 60

    idx = group.index[0]
    hours_all, minutes_all = divmod(int(total_minutes_all), 60)
    df.at[idx, 'ชื่อ(รวม)'] = name
    df.at[idx, 'เวลารวมทั้งหมด'] = f"{hours_all:02}:{minutes_all:02}"

    if any(p in name_str for p in target_people):
        h_p, m_p = divmod(int(press_minutes), 60)
        df.at[idx, 'เวลากดแผ่น'] = f"{h_p:02}:{m_p:02}"
        pack_minutes = max(total_minutes_all - press_minutes, 0)
        h_k, m_k = divmod(int(pack_minutes), 60)
        df.at[idx, 'เวลาห้องแพ็ค'] = f"{h_k:02}:{m_k:02}"
    else:
        df.at[idx, 'เวลาห้องแพ็ค'] = f"{hours_all:02}:{minutes_all:02}"

# ลบคอลัมน์ช่วย
df = df.drop(columns=['วันที่'])

# === STEP 4: เขียนไฟล์ Excel ชั่วคราว ===
output_path = f"excel/calculater_{pre_output}.xlsx"
df.to_excel(output_path, index=False)

# === STEP 5: ไฮไลท์สีคอลัมน์ G, H, I ===
wb = load_workbook(output_path)
ws = wb.active
ws.title = "time"

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # G
orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # H
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # I

for row in range(2, ws.max_row + 1):
    if ws.cell(row, 7).value:
        ws.cell(row, 7).fill = green_fill
    if ws.cell(row, 8).value:
        ws.cell(row, 8).fill = orange_fill
    if ws.cell(row, 9).value:
        ws.cell(row, 9).fill = yellow_fill

# ปรับความกว้างคอลัมน์
for i, width in enumerate([10, 10, 20, 10, 10, 10, 10, 10, 10, 10], start=1):
    col_letter = chr(64 + i)
    ws.column_dimensions[col_letter].width = width

# === STEP 6–8: สร้างชีต "คำนวณ" พร้อมเบิก/จ่ายจริง และ SUM ได้ ===
summary_headers = [
    "ชื่อพนักงาน", "ชั่วโมงห้องแพ็ค", "อัตรา/ชม (แพ็ค)", "รวมเงิน (แพ็ค)",
    "ชั่วโมงกดแผ่น", "อัตรา/ชม (กด)", "รวมเงิน (กด)",
    "รวมจ่ายทั้งหมด", "เบิก", "ประกันสังคม", "จ่ายจริง"
]
ws_summary = wb.create_sheet(title="คำนวณ")
ws_summary.append(summary_headers)

summary_rows = []
for _, row in df.iterrows():
    name = str(row['ชื่อ(รวม)']).strip()
    if not name:
        continue

    pack_time = row['เวลาห้องแพ็ค']
    pack_hours = int(pack_time.split(":")[0]) + int(pack_time.split(":")[1]) / 60

    press_time = row['เวลากดแผ่น']
    press_hours = 0
    if press_time:
        press_hours = int(press_time.split(":")[0]) + int(press_time.split(":")[1]) / 60

    pay_rate = data.get(name, 0)
    price_kod = data['price_kod']
    total_pack_pay = pack_hours * pay_rate
    total_press_pay = press_hours * price_kod
    total_pay = total_pack_pay + total_press_pay

    summary_rows.append([
        name,
        round(pack_hours, 2),
        pay_rate,
        round(total_pack_pay, 2),
        round(press_hours, 2),
        price_kod,
        round(total_press_pay, 2),
        round(total_pay, 2),
        "",  # เบิก
        "",  # ประกันสังคม
        "",  # จ่ายจริง
    ])

for row_data in summary_rows:
    row_data[8] = 0  # ตั้งค่าเริ่มต้นเป็น 0 สำหรับ "เบิก"
    row_data[9] = 0  # ตั้งค่าเริ่มต้นเป็น 0 สำหรับ "ประกันสังคม"

    ws_summary.append(row_data)

# เพิ่มสูตร SUM ที่คอลัมน์รวมจ่ายทั้งหมด (H)
sum_row_idx = len(summary_rows) + 2
ws_summary[f"H{sum_row_idx}"] = f"=SUM(H2:H{sum_row_idx - 1})"
ws_summary[f"H{sum_row_idx}"].alignment = Alignment(horizontal='right')

# === เพิ่มสูตร Excel ที่คอลัมน์ "จ่ายจริง" (K) ===
for row in range(2, ws_summary.max_row + 1):
    ws_summary[f"K{row}"] = f"=ROUND(H{row}-I{row}-J{row}, 0)"  # จ่ายจริง = รวมจ่ายทั้งหมด - เบิก - ประกันสังคม
    ws_summary[f"K{row}"].alignment = Alignment(horizontal='right')
    ws_summary[f"K{row}"].number_format = "#,##0"  # กำหนดให้แสดงคอมม่า

# เพิ่ม Table
table = Table(displayName="PaymentSummary", ref=f"A1:K{len(summary_rows)+1}")
style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
table.tableStyleInfo = style
ws_summary.add_table(table)

# ปรับข้อความชิดขวาคอลัมน์รวมจ่าย
for row in range(2, ws_summary.max_row + 1):
    ws_summary.cell(row=row, column=8).alignment = Alignment(horizontal='right')

# ปรับความกว้างคอลัมน์ A–J
for i, width in enumerate([15] * 11, start=1):
    col_letter = chr(64 + i)
    ws_summary.column_dimensions[col_letter].width = width

# === STEP 9: บันทึกไฟล์ ===
wb.save(output_path)
print(f"✅ บันทึกไฟล์ไฮไลท์แล้วที่: {output_path}")